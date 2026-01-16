from io import BytesIO
from datetime import date as date_type
from datetime import datetime as datetime_type

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

RS_FORMAT = '"Rs." #,##0.00;-"Rs." #,##0.00'
DATE_FORMAT = "MMM DD, YYYY"
MONTH_FORMAT = "MMM YYYY"


def export_workbook_response(workbook, filename):
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def build_daily_workbook(daily_records):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Daily"
    _write_daily_sheet(worksheet, daily_records)
    return workbook


def build_monthly_workbook(monthly_records, daily_records=None):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Monthly"
    _write_monthly_sheet(worksheet, monthly_records)
    if daily_records is not None:
        daily_sheet = workbook.create_sheet("Daily")
        _write_daily_sheet(daily_sheet, daily_records)
    return workbook


def _write_daily_sheet(worksheet, daily_records):
    headers = ["Date", "Sales", "Purchases", "Returns", "Expenses", "Net"]
    _append_header(worksheet, headers)
    for record in daily_records:
        worksheet.append(
            [
                record.get("day"),
                record.get("sales_total", 0),
                record.get("purchases_total", 0),
                record.get("returns_total", 0),
                record.get("expenses_total", 0),
                record.get("net_total", 0),
            ]
        )
    totals = _sum_fields(
        daily_records,
        ("sales_total", "purchases_total", "returns_total", "expenses_total", "net_total"),
    )
    worksheet.append(
        [
            "Total",
            totals["sales_total"],
            totals["purchases_total"],
            totals["returns_total"],
            totals["expenses_total"],
            totals["net_total"],
        ]
    )
    _finalize_sheet(worksheet, DATE_FORMAT)


def _write_monthly_sheet(worksheet, monthly_records):
    headers = ["Month", "Sales", "Purchases", "Returns", "Expenses", "Net"]
    _append_header(worksheet, headers)
    for record in monthly_records:
        worksheet.append(
            [
                record.get("month"),
                record.get("sales_total", 0),
                record.get("purchases_total", 0),
                record.get("returns_total", 0),
                record.get("expenses_total", 0),
                record.get("net_total", 0),
            ]
        )
    totals = _sum_fields(
        monthly_records,
        ("sales_total", "purchases_total", "returns_total", "expenses_total", "net_total"),
    )
    worksheet.append(
        [
            "Total",
            totals["sales_total"],
            totals["purchases_total"],
            totals["returns_total"],
            totals["expenses_total"],
            totals["net_total"],
        ]
    )
    _finalize_sheet(worksheet, MONTH_FORMAT)


def _append_header(worksheet, headers):
    worksheet.append(headers)
    header_font = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _sum_fields(records, fields):
    totals = {}
    for field in fields:
        totals[field] = sum((row.get(field) or 0) for row in records)
    return totals


def _finalize_sheet(worksheet, date_format):
    worksheet.freeze_panes = "A2"
    max_row = worksheet.max_row
    if max_row >= 2:
        for row in worksheet.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=6):
            for cell in row:
                cell.number_format = RS_FORMAT
                cell.alignment = Alignment(horizontal="right")
        for row in worksheet.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=1):
            cell = row[0]
            if isinstance(cell.value, (date_type, datetime_type)):
                cell.number_format = date_format
                cell.alignment = Alignment(horizontal="left")
        for cell in worksheet[max_row]:
            cell.font = Font(bold=True)
    _set_column_widths(worksheet)


def _set_column_widths(worksheet):
    widths = [18, 16, 16, 16, 16, 16]
    for idx, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(idx)].width = width
